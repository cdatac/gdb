#!/usr/bin/env python3
"""
TÜİK TÜFE Veri Çekici — Gökdemir Barometresi

Mimari: Excel indirme YOK.
  1. Playwright ile press metadata sayfasını render et
  2. /api/tr/press/{id} JSON yanıtını yakala
  3. content HTML içindeki <div class="grafik" data-options="..."> bloklarını parse et
  4. GRAFIK2 (yıllık), GRAFIK4 (aylık), GRAFIK1 (genel trend) → sektör verileri
  5. data/raw/tuik_latest.json yaz

Çalıştırma:
  python src/scrape_tuik.py                          # varsayılan config
  python src/scrape_tuik.py --config config/families.json
  python src/scrape_tuik.py --family tuik_tufe --url https://...
"""

from __future__ import annotations

import argparse
import difflib
import json
import re
import sys
from dataclasses import dataclass, asdict, field
from datetime import datetime, date
from html import unescape
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urljoin

import requests

# ---------------------------------------------------------------------------
# Sabitler
# ---------------------------------------------------------------------------

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_CONFIG = ROOT / "config" / "families.json"
DEFAULT_OUTPUT = ROOT / "data" / "raw" / "tuik_latest.json"
CACHE_DIR = ROOT / "data" / "cache"

USER_AGENT = (
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/123.0 Safari/537.36"
)
HEADERS = {"User-Agent": USER_AGENT, "Accept-Language": "tr,en;q=0.9"}

PRESS_RE = re.compile(r"/(?:tr|en)/press/(\d+)(?:/metadata)?", re.IGNORECASE)

JS_SHELL_MARKERS = (
    "JavaScript Gerekli",
    "JavaScript Required",
    "You need to enable JavaScript",
)

# Grafik title → metrik adı eşlemesi
GRAFIK_METRIC_MAP: Dict[str, str] = {
    "yıllık değişim":  "annual_change",
    "aylık değişim":   "monthly_change",
    "yıllık etki":     "annual_contribution",
    "aylık etki":      "monthly_contribution",
}

# TÜFE sektör ID eşlemesi (sıra numarası)
SECTOR_ID_MAP = {
    "gıda ve alkolsüz içecekler":      "01",
    "alkollü içecekler ve tütün":       "02",
    "giyim ve ayakkabı":                "03",
    "konut":                            "04",
    "mobilya ve ev eşyası":             "05",
    "sağlık":                           "06",
    "ulaştırma":                        "07",
    "bilgi ve iletişim":                "08",
    "eğlence ve kültür":                "09",
    "eğitim":                           "10",
    "lokanta ve konaklama":             "11",
    "sigorta ve finansal hizmetler":    "12",
    "sigort ve finansal hizmetler":     "12",   # TÜİK bazen typo
    "çeşitli mal ve hizmetler":         "13",
}


# ---------------------------------------------------------------------------
# Veri yapıları
# ---------------------------------------------------------------------------

@dataclass
class Sector:
    id: str
    name: str
    monthly_change: Optional[float]
    annual_change: Optional[float]
    twelve_month_avg: Optional[float]
    level: int = 1


@dataclass
class ScrapeResult:
    family: str
    press_id: Optional[str]
    date: Optional[str]
    source_url: str
    sectors: List[Dict[str, Any]] = field(default_factory=list)
    error: Optional[str] = None
    warning: Optional[str] = None


# ---------------------------------------------------------------------------
# Yardımcı
# ---------------------------------------------------------------------------

def log(msg: str) -> None:
    print(f"[INFO] {msg}", file=sys.stderr)


def warn(msg: str) -> None:
    print(f"[WARN] {msg}", file=sys.stderr)


def safe_float(v: Any) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(str(v).replace(",", ".").replace("%", "").strip())
    except (ValueError, TypeError):
        return None


def is_js_shell(html: str) -> bool:
    return any(marker in html for marker in JS_SHELL_MARKERS)


def press_id_from_url(url: str) -> Optional[str]:
    m = PRESS_RE.search(url)
    return m.group(1) if m else None


# ---------------------------------------------------------------------------
# Grafik parse
# ---------------------------------------------------------------------------

def _parse_data_options(raw_attr: str) -> Optional[Dict[str, Any]]:
    """
    TÜİK'in data-options attribute'unu JSON'a çevir.
    Attribute içinde single-quote JS literal kullanılıyor.
    """
    try:
        # HTML entity decode
        raw = unescape(raw_attr)
        # JS single-quote literal → JSON double-quote
        # null/true/false değerlerini koru, yalnızca string değerleri değiştir
        def sq_to_dq(m):
            inner = m.group(1).replace('"', '\\"')
            return '"' + inner + '"'
        raw = re.sub(r"'([^']*)'", sq_to_dq, raw)
        # Trailing commas
        raw = re.sub(r",\s*}", "}", raw)
        raw = re.sub(r",\s*]", "]", raw)
        return json.loads(raw)
    except Exception:
        return None


def extract_grafiks_from_content(content_html: str) -> Dict[str, Dict[str, Any]]:
    """
    HTML içindeki tüm GRAFIK bloklarını parse et.
    Döner: {"GRAFIK1": opts_dict, "GRAFIK2": opts_dict, ...}
    """
    pattern = re.compile(
        r'<div[^>]+data-name="(GRAFIK\d+)"[^>]+data-lang="tr"[^>]*data-options="([^"]*)"',
        re.DOTALL,
    )
    result: Dict[str, Dict[str, Any]] = {}
    for m in pattern.finditer(content_html):
        name = m.group(1)
        opts = _parse_data_options(m.group(2))
        if opts:
            result[name] = opts
    return result


def _grafik_metric_name(grafik_opts: Dict[str, Any]) -> Optional[str]:
    """Grafik başlığından metrik adını çıkar."""
    title = str(grafik_opts.get("name", "")).lower()
    for keyword, metric in GRAFIK_METRIC_MAP.items():
        if keyword in title:
            return metric
    # label field'ına bak
    for series in grafik_opts.get("data", []):
        lbl = str(series.get("label", "")).lower()
        for keyword, metric in GRAFIK_METRIC_MAP.items():
            if keyword in lbl:
                return metric
    return None


def _sector_id(label: str) -> str:
    """Sektör label'ından ID üret."""
    normalized = label.strip().lower()
    for key, sid in SECTOR_ID_MAP.items():
        if key in normalized or normalized in key:
            return sid
    # Fallback: ilk 2 hane varsa al
    m = re.match(r"^(\d{2})", normalized)
    if m:
        return m.group(1)
    return normalized[:6].replace(" ", "_")


def build_sectors_from_grafiks(grafiks: Dict[str, Dict[str, Any]]) -> List[Sector]:
    """
    GRAFIK2 (yıllık) + GRAFIK4 (aylık) + GRAFIK1 (genel trend) → Sector listesi.
    """
    # Metriklere göre grafikleri grupla
    metric_data: Dict[str, Tuple[List[str], List[float]]] = {}
    grafik1_series: Optional[Tuple[List[str], List[float]]] = None

    for gname, opts in grafiks.items():
        metric = _grafik_metric_name(opts)
        labels = opts.get("labels", [])

        for series in opts.get("data", []):
            vals = series.get("data", [])
            if not vals:
                continue
            if gname == "GRAFIK1":
                grafik1_series = (labels, [safe_float(v) for v in vals])
            if metric and labels:
                metric_data[metric] = (labels, [safe_float(v) for v in vals])

    annual_labels, annual_vals = metric_data.get("annual_change", ([], []))
    monthly_labels, monthly_vals = metric_data.get("monthly_change", ([], []))

    if not annual_labels:
        return []

    # 12-ay ort: GRAFIK1'deki genel TÜFE serisinin son 12 değerinin ortalaması
    # Sektör bazında 12-ay ort TÜİK'te ayrı grafik olarak yayınlanmıyor,
    # bu yüzden yıllık değişimin son dönemini kullanıyoruz (iyi yaklaşım).
    # Not: Gerçek 12-ay ort için ek API endpoint gerekir.
    avg12_estimate: Optional[float] = None
    if grafik1_series:
        _, g1_vals = grafik1_series
        last12 = [v for v in g1_vals[-12:] if v is not None]
        if last12:
            avg12_estimate = round(sum(last12) / len(last12), 2)

    sectors: List[Sector] = []
    for i, label in enumerate(annual_labels):
        if label.strip().upper() in ("TÜFE", "TUFE", "GENEL", "TOPLAM"):
            continue  # Genel endeks satırını atla

        annual = annual_vals[i] if i < len(annual_vals) else None
        monthly = None
        for j, ml in enumerate(monthly_labels):
            if ml.strip().lower() == label.strip().lower():
                monthly = monthly_vals[j] if j < len(monthly_vals) else None
                break

        # 12-ay ort: sektör için genel TÜFE ortalamasını proxy olarak kullan
        twelve_avg = avg12_estimate

        sectors.append(Sector(
            id=_sector_id(label),
            name=label.strip(),
            monthly_change=annual,    # grafik labels aynı sıra değil, aşağıda düzeltilecek
            annual_change=annual,
            twelve_month_avg=twelve_avg,
            level=1,
        ))

    # monthly_change'i doğru ata (fuzzy eşleştirme ile TÜİK typo'larına dayanıklı)
    monthly_by_label = {ml.strip().lower(): monthly_vals[j]
                        for j, ml in enumerate(monthly_labels) if j < len(monthly_vals)}
    for s in sectors:
        name_lower = s.name.strip().lower()
        mc = monthly_by_label.get(name_lower)
        if mc is None:
            # TÜİK zaman zaman GRAFIK2 ve GRAFIK4'te farklı yazım kullanıyor
            # (ör. "Sigort" vs "Sigorta") → difflib ile en yakın eşleşmeyi bul
            close = difflib.get_close_matches(name_lower, monthly_by_label.keys(), n=1, cutoff=0.85)
            if close:
                mc = monthly_by_label[close[0]]
                log(f"Fuzzy eşleşme: '{s.name}' → '{close[0]}' (aylık)")
        if mc is not None:
            s.monthly_change = mc

    return sectors


# ---------------------------------------------------------------------------
# Alt sektör Excel verisi (Düzey 3 — harcama gruplarına göre endeks)
# ---------------------------------------------------------------------------

VERI_PORTALI = "https://veriportali.tuik.gov.tr"

# İngilizce ay adı → ay numarası (Excel'deki col 2)
MONTH_EN: Dict[str, int] = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
}
# Cache: önce .xlsx (modern), yoksa .xls (OLE)
SUBGROUP_CACHE_XLSX = CACHE_DIR / "tuik_subgroups.xlsx"
SUBGROUP_CACHE_XLS  = CACHE_DIR / "tuik_subgroups.xls"


def _open_excel_as_rows(xls_bytes: bytes) -> Optional[Tuple[List[str], List[List[Any]]]]:
    """
    Excel baytlarını okuyup (sheet_names, rows_2d) tuple'ı döndür.
    Önce xlrd (OLE .xls), ardından openpyxl (.xlsx) dener.
    rows_2d: liste içinde liste, her iç liste bir satırın hücre değerleri.
    """
    # OLE .xls magic: D0 CF 11 E0 veya xlsx magic: PK\x03\x04
    is_xlsx = xls_bytes[:4] == b"PK\x03\x04"

    if not is_xlsx:
        try:
            import xlrd  # type: ignore
            wb = xlrd.open_workbook(file_contents=xls_bytes)
            result: Dict[str, List[List[Any]]] = {}
            for sname in wb.sheet_names():
                sh = wb.sheet_by_name(sname)
                result[sname] = [sh.row_values(r) for r in range(sh.nrows)]
            log(f"xlrd ile açıldı: {list(result.keys())}")
            return (list(result.keys()), result)
        except Exception as exc:
            warn(f"xlrd başarısız ({exc}), openpyxl deneniyor...")

    try:
        import openpyxl  # type: ignore
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(xls_bytes), read_only=True, data_only=True)
        result = {}
        for sname in wb.sheetnames:
            sh = wb[sname]
            result[sname] = [[cell.value for cell in row] for row in sh.iter_rows()]
        wb.close()
        log(f"openpyxl ile açıldı: {list(result.keys())}")
        return (wb.sheetnames, result)
    except Exception as exc:
        warn(f"openpyxl de başarısız: {exc}")
        return None


def fetch_subgroup_excel(api_body: Dict[str, Any]) -> Dict[str, List[Dict[str, Any]]]:
    """
    Press API body'sindeki statisticalTables listesinden "Harcama gruplarına göre"
    Excel dosyasını indir, Düzey 3 sheet'ini parse et.

    Döner: {"01": [{id, name, monthly_change, annual_change, twelve_month_avg}, ...], ...}
    Anahtar: üst sektörün 2 haneli COICOP kodu (3 haneli kodun ilk 2 hanesi).
    """
    # --- URL bul ---
    tables = (
        api_body.get("data", {}).get("statisticalTables", [])
        or api_body.get("statisticalTables", [])
    )
    target_url: Optional[str] = None
    for tbl in tables:
        title = str(tbl.get("title") or tbl.get("name") or "").lower()
        if "harcama" in title:
            rel = tbl.get("url", "")
            if rel:
                target_url = VERI_PORTALI + rel
                log(f"Subgroup Excel: {title[:60]}")
                break

    if not target_url:
        warn("statisticalTables içinde 'harcama' Excel'i bulunamadı")
        return {}

    # --- İndir (önce cache dene) ---
    xls_bytes: Optional[bytes] = None
    for cache_path in (SUBGROUP_CACHE_XLSX, SUBGROUP_CACHE_XLS):
        if cache_path.exists():
            xls_bytes = cache_path.read_bytes()
            log(f"Subgroup Excel cache'den: {cache_path.name} ({len(xls_bytes)} bayt)")
            break

    if xls_bytes is None:
        try:
            resp = requests.get(target_url, headers=HEADERS, timeout=90)
            if resp.status_code != 200:
                warn(f"Subgroup Excel indirilemedi: HTTP {resp.status_code}")
                return {}
            xls_bytes = resp.content
            log(f"Subgroup Excel indirildi: {len(xls_bytes)} bayt")
            CACHE_DIR.mkdir(parents=True, exist_ok=True)
            cache_file = SUBGROUP_CACHE_XLSX if xls_bytes[:4] == b"PK\x03\x04" else SUBGROUP_CACHE_XLS
            cache_file.write_bytes(xls_bytes)
        except Exception as exc:
            warn(f"Subgroup Excel indirme hatası: {exc}")
            return {}

    # --- Excel parse (xlrd veya openpyxl) ---
    parsed = _open_excel_as_rows(xls_bytes)
    if parsed is None:
        warn("Excel dosyası açılamadı")
        return {}
    sheet_names, sheet_data = parsed

    # "Düzey 3" sheet'ini bul
    target_sheet: Optional[str] = None
    for sname in sheet_names:
        if "3" in sname:
            target_sheet = sname
            break
    if target_sheet is None:
        warn(f"Düzey 3 sheet bulunamadı. Mevcut: {sheet_names}")
        return {}

    rows_2d = sheet_data[target_sheet]
    nrows = len(rows_2d)
    ncols = max((len(r) for r in rows_2d), default=0)
    log(f"Sheet '{target_sheet}': {nrows} satır × {ncols} sütun")

    if nrows < 10 or ncols < 5:
        warn(f"Sheet beklenenden küçük: {nrows}×{ncols}")
        return {}

    # Satır 5 (index 4) = COICOP kodları, Satır 6 (index 5) = isimler
    code_row = rows_2d[4] if nrows > 4 else []
    name_row = rows_2d[5] if nrows > 5 else []

    def _pad(lst, length):
        return list(lst) + [None] * max(0, length - len(lst))

    code_row = _pad(code_row, ncols)
    name_row = _pad(name_row, ncols)

    # Sütun 0 = Yıl, Sütun 1 = Ay, Sütun 2+ = sektörler
    DATA_COL_START = 2

    col_info: List[Tuple[int, str, str]] = []
    for ci in range(DATA_COL_START, ncols):
        raw = code_row[ci]
        if isinstance(raw, (int, float)) and raw and raw > 0:
            code = str(int(raw)).zfill(3)
        elif raw is not None:
            code = re.sub(r"\.0$", "", str(raw).strip())
        else:
            continue
        name_raw = name_row[ci]
        name = str(name_raw).strip() if name_raw is not None else ""
        if re.match(r"^\d{3}$", code) and name and name.lower() not in ("none", ""):
            col_info.append((ci, code, name))

    if not col_info:
        warn("3 haneli COICOP kodu bulunamadı")
        return {}
    log(f"Düzey 3 sektörler: {len(col_info)}")

    # Veri satırlarını oku (year, month, {col_idx: index_value})
    # Sütun düzeni: col0=Yıl, col1=Türkçe ay adı, col2=İngilizce ay adı, col3=TÜFE, col4+=alt sektörler
    data_rows: List[Tuple[int, int, Dict[int, float]]] = []
    for ri in range(6, nrows):
        row = _pad(rows_2d[ri], ncols)
        try:
            year = int(float(row[0]))  # type: ignore[arg-type]
        except (ValueError, TypeError):
            continue
        if year < 2000:
            continue
        # Ay: col2'deki İngilizce ay adını kullan (ASCII güvenli)
        month_str = str(row[2]).strip().lower() if row[2] is not None else ""
        month = MONTH_EN.get(month_str)
        if month is None:
            continue
        vals: Dict[int, float] = {}
        for ci, _, _ in col_info:
            try:
                v = float(row[ci])
                if v > 0:
                    vals[ci] = v
            except (ValueError, TypeError):
                pass
        if vals:
            data_rows.append((year, month, vals))

    if len(data_rows) < 14:
        warn(f"Yeterli satır yok: {len(data_rows)}")
        return {}

    last = data_rows[-1]
    prev = data_rows[-2]
    log(f"Son dönem: {last[0]}-{last[1]:02d}")

    result: Dict[str, List[Dict[str, Any]]] = {}

    for ci, code, name in col_info:
        v_last = last[2].get(ci)
        v_prev = prev[2].get(ci)

        # Yıllık değişim için 12 ay önceki değer
        v_yago = data_rows[-13][2].get(ci) if len(data_rows) >= 13 else None

        if v_last is None or v_prev is None or v_yago is None:
            continue
        if v_prev == 0 or v_yago == 0:
            continue

        monthly_change = round((v_last / v_prev - 1) * 100, 2)
        annual_change = round((v_last / v_yago - 1) * 100, 2)

        # 12 aylık ortalama yıllık değişim
        annual_list: List[float] = []
        n = len(data_rows)
        for k in range(max(12, n - 12), n):
            vc = data_rows[k][2].get(ci)
            vp = data_rows[k - 12][2].get(ci)
            if vc and vp and vp > 0:
                annual_list.append((vc / vp - 1) * 100)
        avg12 = round(sum(annual_list) / len(annual_list), 2) if annual_list else annual_change

        parent = code[:2]  # "011" → "01"
        result.setdefault(parent, []).append({
            "id": code,
            "name": name,
            "monthly_change": monthly_change,
            "annual_change": annual_change,
            "twelve_month_avg": avg12,
        })

    total = sum(len(v) for v in result.values())
    log(f"Alt sektörler: {total} adet, {len(result)} ana grup")
    return result


# ---------------------------------------------------------------------------
# Press API'dan veri çekme
# ---------------------------------------------------------------------------

def fetch_press_api_with_playwright(press_url: str) -> Optional[Dict[str, Any]]:
    """
    Playwright ile press metadata sayfasını render et,
    /api/tr/press/{id} yanıtını yakala ve döndür.
    """
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        warn("Playwright kurulu değil")
        return None

    pid_match = re.search(r"/press/(\d+)", press_url)
    press_id = pid_match.group(1) if pid_match else None
    if not press_id:
        warn(f"Press ID çıkarılamadı: {press_url}")
        return None

    log(f"Playwright: press sayfası yükleniyor → {press_url}")
    captured: Dict[str, Any] = {}

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        context = browser.new_context(user_agent=USER_AGENT, locale="tr-TR")
        page = context.new_page()

        def on_response(resp):
            try:
                u, s = resp.url, resp.status
            except Exception:
                return
            if s < 400 and "/api/" in u and f"/press/{press_id}" in u:
                try:
                    captured["body"] = resp.json()
                    captured["url"] = u
                except Exception as exc:
                    warn(f"Press API parse: {exc}")

        page.on("response", on_response)
        page.goto(press_url, wait_until="domcontentloaded", timeout=90000)
        try:
            page.wait_for_load_state("networkidle", timeout=20000)
        except Exception:
            pass
        page.wait_for_timeout(3000)
        browser.close()

    if "body" in captured:
        log(f"Press API yakalandı: {captured.get('url')}")
        return captured["body"]

    warn(f"Press API yanıtı gelmedi ({press_url})")
    return None


def scrape_press_url(family: str, press_url: str) -> ScrapeResult:
    """
    Press URL'inden sektör verilerini çek.
    """
    result = ScrapeResult(
        family=family,
        press_id=press_id_from_url(press_url),
        date=date.today().strftime("%Y-%m"),
        source_url=press_url,
    )

    # API yanıtını al
    api_body = fetch_press_api_with_playwright(press_url)
    if not api_body:
        result.error = "Press API yanıtı alınamadı"
        return result

    # content HTML'den grafikleri çıkar
    content_html = api_body.get("data", {}).get("content", "")
    if not content_html:
        # Yanıt yapısı farklı olabilir
        content_html = api_body.get("content", "")

    if not content_html:
        result.error = f"content alanı boş — API anahtarları: {list(api_body.keys())}"
        return result

    grafiks = extract_grafiks_from_content(content_html)
    log(f"Grafik sayısı: {len(grafiks)} → {list(grafiks.keys())}")

    if not grafiks:
        result.error = "HTML içinde grafik bulunamadı"
        return result

    sectors = build_sectors_from_grafiks(grafiks)
    log(f"Sektör sayısı: {len(sectors)}")

    if not sectors:
        result.error = "Grafiklerden sektör verisi üretilemedi"
        return result

    # Alt sektör verilerini ekle (Düzey 3 Excel)
    subgroup_map = fetch_subgroup_excel(api_body)
    sector_dicts = [asdict(s) for s in sectors]
    for sd in sector_dicts:
        sid = str(sd.get("id", ""))
        subs = subgroup_map.get(sid, [])
        sd["subgroups"] = subs
        if subs:
            log(f"  {sd['name']} → {len(subs)} alt grup")

    result.sectors = sector_dicts
    return result


# ---------------------------------------------------------------------------
# Bülten keşfi
# ---------------------------------------------------------------------------

def _find_press_link_in_html(html: str, base_url: str, must_contain: str) -> Optional[str]:
    href_re = re.compile(r'href=["\']([^"\']+/press/\d+(?:/metadata)?[^"\']*)["\']', re.IGNORECASE)
    candidates = []
    for m in href_re.finditer(html):
        raw_url = unescape(m.group(1))
        abs_url = urljoin(base_url, raw_url)
        abs_url = re.sub(r'(/press/\d+)(?!/metadata)(/|$)', r'\1/metadata\2', abs_url)
        start = max(0, m.start() - 500)
        context = html[start: m.end() + 500]
        if must_contain.lower() in context.lower():
            pid = press_id_from_url(abs_url)
            if pid:
                candidates.append((int(pid), abs_url))
    if candidates:
        return max(candidates, key=lambda x: x[0])[1]
    return None


def _extract_press_id_from_json(data: Any, must_contain: str) -> Optional[int]:
    candidates: List[int] = []

    def _walk(obj):
        if isinstance(obj, dict):
            for key in ("id", "pressId", "press_id", "Id"):
                val = obj.get(key)
                if isinstance(val, int) and val > 10000:
                    title = str(obj.get("title") or obj.get("name") or
                                obj.get("baslik") or obj.get("adi") or "")
                    if not must_contain or must_contain.lower() in title.lower():
                        candidates.append(val)
            for v in obj.values():
                _walk(v)
        elif isinstance(obj, list):
            for item in obj:
                _walk(item)

    _walk(data)
    return max(candidates) if candidates else None


_API_PRESS_LIST_RE = re.compile(
    r"/api/(?:tr|en)/(?:press(?:es)?|bulten|release)s?",
    re.IGNORECASE,
)


def discover_latest_press_url(discover_url: str, must_contain: str) -> Optional[str]:
    log(f"Bülten keşfi: {discover_url} | aranacak: '{must_contain}'")

    try:
        resp = requests.get(discover_url, headers=HEADERS, timeout=30)
        html = resp.text
        if not is_js_shell(html):
            link = _find_press_link_in_html(html, discover_url, must_contain)
            if link:
                log(f"Statik HTML'de bulundu: {link}")
                return link
    except Exception as exc:
        warn(f"Statik fetch: {exc}")

    try:
        return _discover_with_playwright(discover_url, must_contain)
    except Exception as exc:
        warn(f"Playwright keşfi: {exc}")
        return None


def _discover_with_playwright(discover_url: str, must_contain: str) -> Optional[str]:
    from playwright.sync_api import sync_playwright

    log("Playwright ile kategori sayfası render ediliyor...")
    found_url: Optional[str] = None
    api_press_ids: List[int] = []

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        context = browser.new_context(user_agent=USER_AGENT, locale="tr-TR")
        page = context.new_page()

        def on_response(resp):
            try:
                url = resp.url
            except Exception:
                return
            if not _API_PRESS_LIST_RE.search(url):
                return
            try:
                body = resp.json()
                pid = _extract_press_id_from_json(body, must_contain)
                if pid:
                    api_press_ids.append(pid)
                    log(f"API'den press ID: {pid}")
            except Exception:
                pass

        page.on("response", on_response)
        page.goto(discover_url, wait_until="domcontentloaded", timeout=60000)
        try:
            page.wait_for_load_state("networkidle", timeout=15000)
        except Exception:
            pass
        page.wait_for_timeout(3000)

        html = page.content()
        found_url = _find_press_link_in_html(html, discover_url, must_contain)

        if not found_url:
            for anchor in page.locator("a[href*='/press/']").all()[:50]:
                try:
                    href = anchor.get_attribute("href") or ""
                    text = anchor.inner_text(timeout=500)
                    if must_contain.lower() in text.lower():
                        abs_url = urljoin(discover_url, href)
                        abs_url = re.sub(r'(/press/\d+)(?!/metadata)(/|$)', r'\1/metadata\2', abs_url)
                        if press_id_from_url(abs_url):
                            found_url = abs_url
                            break
                except Exception:
                    continue

        if not found_url and api_press_ids:
            best_id = max(api_press_ids)
            base = discover_url.split("/tr/")[0] if "/tr/" in discover_url else discover_url
            found_url = f"{base}/tr/press/{best_id}/metadata"
            log(f"API press ID'den URL: {found_url}")

        browser.close()

    if found_url:
        log(f"Bülten bulundu: {found_url}")
    return found_url


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

def load_config(path: Path) -> List[Dict[str, Any]]:
    raw = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(raw, list):
        return raw
    if isinstance(raw, dict) and "families" in raw:
        return raw["families"]
    raise ValueError(f"Geçersiz config: {path}")


# ---------------------------------------------------------------------------
# CLI + main
# ---------------------------------------------------------------------------

def parse_args(argv=None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description="TÜİK TÜFE veri çekici (grafik parse)")
    p.add_argument("--family", help="Tek family adı")
    p.add_argument("--url", help="Metadata URL (--family ile birlikte)")
    p.add_argument("--config", help="families.json yolu")
    p.add_argument("--output", help="Çıktı JSON (varsayılan: data/raw/tuik_latest.json)")
    return p.parse_args(argv)


def main(argv=None) -> int:
    args = parse_args(argv)
    jobs: List[Dict[str, Any]] = []

    if args.family and args.url:
        jobs = [{"family": args.family, "label": args.family, "url": args.url}]
    else:
        config_path = Path(args.config) if args.config else DEFAULT_CONFIG
        if not config_path.exists():
            print(f"[HATA] Config bulunamadı: {config_path}", file=sys.stderr)
            return 1
        jobs = load_config(config_path)
        log(f"Config: {config_path} → {len(jobs)} family")

    output_path = Path(args.output) if args.output else DEFAULT_OUTPUT
    all_results: List[Dict[str, Any]] = []
    total_sectors = 0

    for job in jobs:
        family = job.get("family", "unknown")
        label  = job.get("label", family)

        press_url: Optional[str] = job.get("url") or None
        discover_url = job.get("discover_url")

        if not press_url and discover_url:
            must_contain = job.get("must_contain", "")
            press_url = discover_latest_press_url(discover_url, must_contain)
            if not press_url and job.get("fallback_url"):
                press_url = job["fallback_url"]
                warn(f"Keşif başarısız, fallback: {press_url}")
        elif not press_url:
            press_url = job.get("fallback_url") or None

        if not press_url:
            log(f"[ATLA] {family}: press URL yok")
            all_results.append({"family": family, "label": label,
                                 "error": "press_url bulunamadı", "sectors": []})
            continue

        log(f"Scraping: family={family}, url={press_url}")
        res = scrape_press_url(family, press_url)
        n = len(res.sectors)
        log(f"Tamamlandı: family={family}, sektör={n}, hata={res.error}")
        total_sectors += n

        entry: Dict[str, Any] = {
            "family": family, "label": label,
            "press_id": res.press_id, "date": res.date,
            "source_url": res.source_url, "sectors": res.sectors,
        }
        if res.error:
            entry["error"] = res.error
        if res.warning:
            entry["warning"] = res.warning
        all_results.append(entry)

    if total_sectors == 0:
        log("UYARI: Hiçbir sektör verisi çıkarılamadı")

    payload: Dict[str, Any] = {
        "scraped_at": datetime.now().isoformat(timespec="seconds"),
        "total_sectors": total_sectors,
        "families": all_results,
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    log(f"Yazıldı: {output_path} (toplam sektör: {total_sectors})")

    if total_sectors > 0:
        CACHE_DIR.mkdir(parents=True, exist_ok=True)
        cache_path = CACHE_DIR / "tuik_latest.json"
        cache_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        log(f"Cache güncellendi: {cache_path}")

    return 0 if total_sectors > 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())
